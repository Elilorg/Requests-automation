import csv 
from time import date 

#detache le destinataire et le moyen de payement à partire de la colonne 
def traitement(text : str) -> (str,str):
    if "ACHAT CB" in text : 
        text = text.replace("ACHAT CB", "")
        destinataire = text[:-43].strip()
        return destinataire.replace('"', "").strip(), "CB"
    
    elif "PRELEVEMENT DE" in text : 
        text = text.replace("PRELEVEMENT DE", "")
        destinataire = text.split("REF")[0]
        return destinataire.replace('"', "").strip(), "PRELEVEMENT"
    
    elif "VIREMENT INSTANTANE A" in text :
        text = text.replace('VIREMENT INSTANTANE A', "")
        destinataire = text
        return destinataire.replace('"', "").strip(), "VIREMENT SORTANT"
    
    elif "VIREMENT INSTANTANE DE" in text : 
        text = text.replace('VIREMENT INSTANTANE DE', "")
        destinataire = text
        return destinataire.strip(), "VIREMENT ENTRANT"
    
    else : return "LIQUIDE", "RETRAIT"


categories = {
            "TRANSPORT" : ['SNCF', 'SNCF INTERNET'], 
            "NOURRITURE" : ["CARREFOUR", 'SC-BOUL CHARRI', 'IZLY SMONEY', 'AUCHAN SUPERMA', 'MAISON ISAAC','L ORIENT EXPRE'],
            "UTC" : ["UTC", "WEEZEVENT", 'BDE UTC PAYUTC'], 
            "Réguliers" : ['EDF clients parti iers', "MATMUT ROUEN"], 
            "AUTRE" : [""]
            }


#donne la catégorie du destinnataire en conaissant le destinataire 
def categorie(destinataire : str) -> str : 
    for i in categories.keys() :
        if destinataire in categories[i] : 
            return i 
    return "AUTRE"

def traitement_de_ligne(row : list[str*3]) -> list[str, int] : 
    destinataire, moyen_de_payement = traitement(row[1])
    row[1] = destinataire
    row.append(moyen_de_payement)

    float_montant = float(row[2].replace(",","."))
    row[2] = float_montant

    row.append(categorie(destinataire))
    return row


with open('relevé.csv') as csvfile : 
    csvreader = csv.reader(csvfile, delimiter=";")

    header = [next(csvreader) for i in range(7)]
    new_file = []
    for row in csvreader : 

        row = traitement_de_ligne(row)
        new_file.append(row)
        print(row)

###Créations des feuilles annexes et des graphes 
import openpyxl

firstline = header[-1]
firstline[1] = "Destinataire"
firstline[2] = "Montant"
firstline += ["moyen de payement","Catégorie"]

wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(firstline) 

 

for i in new_file : 
    if i[2] < 0 :
        sheet.append(i)


"""
#graphe pour je ne sais quoi 
chart_ref = openpyxl.chart.Reference(sheet,min_col=3,max_col=3,min_row=1,max_row=sheet.max_row )
chart_series = openpyxl.chart.Series(chart_ref)
chart = openpyxl.chart.LineChart()
chart.append(chart_series)
sheet.add_chart(chart,"F5")
"""


#somme des dépenses par date 
wb.create_sheet(title="Par jour")
sheet2 = wb["Par jour"]

#trouver comment ajouter toutes les dates du mois
"""
1. créer une liste de toutes les dates du moi donné
2. faire la meme fonction pour toutes les dates """

dates = sorted(list(set([i[0] for i in new_file])))
for i in dates :
    sheet2.append([i, abs(sum([n[2] for n in new_file if n[2] < 0 and n[0] == i]))])
    

#graphe par date 

chart_ref = openpyxl.chart.Reference(sheet2,min_col=2,max_col=3,min_row=sheet2.min_row,max_row=sheet2.max_row )
chart_series = openpyxl.chart.Series(chart_ref)
chart = openpyxl.chart.LineChart()
chart.append(chart_series)
sheet2.add_chart(chart,"F5")

#Dépenses par catégories

wb.create_sheet(title="Par Catégorie")
sheet3 = wb["Par Catégorie"]


cat = categories.keys()
sheet3.append(list(cat))
sheet3.append( [sum([abs(n[2]) for n in new_file if n[2]<0 and n[4] == i]) for i in cat ])

#graphe par catégorie (trouver comment mettre les catégories en abcisse)
chart_ref = openpyxl.chart.Reference(sheet3,min_col=sheet.min_column,max_col=sheet3.max_column,min_row=2,max_row=sheet3.max_row )
chart_series = openpyxl.chart.Series(chart_ref)
chart = openpyxl.chart.BarChart()
chart.append(chart_series)
sheet3.add_chart(chart,"F5")

wb.save('relevé.xlsx')           
